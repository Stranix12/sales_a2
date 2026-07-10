"""Exportación individual (PDF/Excel) de un registro de compra.

Complementa a ``billing.mixins.ExportListMixin``: aquel exporta un *listado*
completo (varias filas, columnas configurables); este módulo exporta **un
único** ``Purchase`` como documento/comprobante (cabecera + líneas de detalle),
que es lo que tiene sentido para un registro individual.

Mismo estándar visual que ``ExportListMixin`` para que todos los documentos
generados por la aplicación se vean consistentes:
  - Excel (``openpyxl``): encabezados en blanco/negrita sobre fondo `#343A40`.
  - PDF (``reportlab``): título + fecha de generación, tabla con cabecera
    `#343A40` y filas alternas grises.
  - Nombre de archivo con marca de tiempo, p. ej. ``Compra_5_20260706_1430.xlsx``.
"""
from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

HEADER_FILL = '343A40'
ACCENT = '33459B'
_ESTADO_COLOR = {'PAGADA': '1F9D57', 'PENDIENTE': 'E0A008'}


def _fecha_corta(d):
    return d.strftime('%d/%m/%Y') if d else '—'


def _section(title, style):
    return [
        Paragraph(title.upper(), style),
        HRFlowable(width='100%', thickness=0.8, color=colors.HexColor(f'#{ACCENT}'),
                  spaceBefore=2, spaceAfter=6),
    ]


def _export_filename(purchase):
    stamp = timezone.localtime().strftime('%Y%m%d_%H%M')
    return f'Compra_{purchase.id}_{stamp}'


def export_purchase_excel(purchase):
    """Genera el .xlsx de una compra individual (cabecera + líneas)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Compra {purchase.id}'[:31]

    bold = Font(bold=True)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type='solid')
    center = Alignment(horizontal='center', vertical='center')

    # --- Cabecera de la compra ---
    ws.append(['Compra #', purchase.id])
    ws.append(['Proveedor', str(purchase.supplier)])
    ws.append(['N° Factura Proveedor', purchase.document_number])
    ws.append(['Fecha', timezone.localtime(purchase.purchase_date).strftime('%d/%m/%Y %H:%M')])
    for row in range(1, 5):
        ws.cell(row=row, column=1).font = bold
    ws.append([])

    # --- Tabla de líneas ---
    header_row = ws.max_row + 1
    headers = ['Producto', 'Cantidad', 'Costo Unitario', 'Subtotal']
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.font, cell.fill, cell.alignment = header_font, header_fill, center

    for detail in purchase.details.all():
        ws.append([detail.product.name, detail.quantity,
                   float(detail.unit_cost), float(detail.subtotal)])

    ws.append(['', '', 'Subtotal', float(purchase.subtotal)])
    ws.append(['', '', 'IVA (15%)', float(purchase.tax)])
    total_row = ws.max_row + 1
    ws.append(['', '', 'TOTAL', float(purchase.total)])
    for col in (3, 4):
        ws.cell(row=total_row, column=col).font = bold

    widths = [28, 12, 16, 14]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{_export_filename(purchase)}.xlsx"'
    return response


def export_purchase_pdf(purchase):
    """Genera el PDF de una compra individual (cabecera + líneas)."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        title=f'Compra #{purchase.id}',
    )
    styles = getSampleStyleSheet()
    normal = styles['Normal']
    label_style = ParagraphStyle('lbl', parent=normal, fontName='Helvetica-Bold')
    section_style = ParagraphStyle('section', parent=normal, fontName='Helvetica-Bold', fontSize=10,
                                   textColor=colors.HexColor(f'#{ACCENT}'))

    elements = [
        Paragraph(f'Compra #{purchase.id}', styles['Title']),
        Paragraph(
            f"Generado: {timezone.localtime().strftime('%d/%m/%Y %H:%M')}", normal),
        Spacer(1, 0.5 * cm),
        Table([
            [Paragraph('Proveedor:', label_style), str(purchase.supplier)],
            [Paragraph('N° Factura Proveedor:', label_style), purchase.document_number],
            [Paragraph('Fecha:', label_style),
             timezone.localtime(purchase.purchase_date).strftime('%d/%m/%Y %H:%M')],
        ], colWidths=[5 * cm, 11 * cm]),
        Spacer(1, 0.6 * cm),
    ]

    # --- Condiciones de pago: contado o crédito (y a cuántos meses) ---
    cuotas = list(purchase.cuotas.all()) if purchase.tipo_pago == 'CREDITO' else []
    elements += _section('Condiciones de pago', section_style)
    if purchase.tipo_pago == 'CREDITO':
        pagadas = sum(1 for c in cuotas if c.estado == 'PAGADA')
        condiciones = [
            [Paragraph('Tipo de pago:', label_style), f'Crédito a {len(cuotas)} meses'],
            [Paragraph('Cuotas pagadas:', label_style), f'{pagadas} de {len(cuotas)}'],
            [Paragraph('Saldo pendiente:', label_style), f'${purchase.saldo}'],
        ]
    else:
        condiciones = [[Paragraph('Tipo de pago:', label_style), purchase.get_tipo_pago_display()]]
    elements += [
        Table(condiciones, colWidths=[3.5 * cm, 12.5 * cm],
              style=TableStyle([('FONTSIZE', (0, 0), (-1, -1), 9)])),
        Spacer(1, 0.5 * cm),
    ]

    elements += _section('Detalle', section_style)
    data = [['Producto', 'Cantidad', 'Costo Unitario', 'Subtotal']]
    for detail in purchase.details.all():
        data.append([
            detail.product.name,
            str(detail.quantity),
            f'${detail.unit_cost}',
            f'${detail.subtotal}',
        ])
    data.append(['', '', 'Subtotal', f'${purchase.subtotal}'])
    data.append(['', '', 'IVA (15%)', f'${purchase.tax}'])
    data.append(['', '', 'TOTAL', f'${purchase.total}'])

    table = Table(data, colWidths=[7 * cm, 3 * cm, 3.5 * cm, 3.5 * cm], repeatRows=1)
    n_lines = len(purchase.details.all())
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f'#{HEADER_FILL}')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, n_lines), [colors.white, colors.HexColor('#F2F2F2')]),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (0, 0), 'CENTER'),
        ('FONTNAME', (2, -1), (-1, -1), 'Helvetica-Bold'),
        ('SPAN', (0, -3), (1, -3)),
        ('SPAN', (0, -2), (1, -2)),
        ('SPAN', (0, -1), (1, -1)),
        ('LINEABOVE', (2, -1), (-1, -1), 1, colors.black),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(table)

    # --- Plan de cuotas (solo si la compra fue a crédito) ---
    if purchase.tipo_pago == 'CREDITO' and cuotas:
        elements.append(Spacer(1, 0.5 * cm))
        elements += _section(f'Plan de cuotas ({len(cuotas)} meses)', section_style)
        cuotas_data = [['#', 'Vencimiento', 'Valor', 'Saldo', 'Estado']]
        for c in cuotas:
            cuotas_data.append([str(c.numero), _fecha_corta(c.fecha_vencimiento),
                               f'${c.valor}', f'${c.saldo}', c.get_estado_display()])
        cuotas_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f'#{HEADER_FILL}')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]
        for i, c in enumerate(cuotas, start=1):
            color = _ESTADO_COLOR.get(c.estado, '5c6474')
            cuotas_style.append(('TEXTCOLOR', (4, i), (4, i), colors.HexColor(f'#{color}')))
            cuotas_style.append(('FONTNAME', (4, i), (4, i), 'Helvetica-Bold'))
        cuotas_table = Table(cuotas_data, colWidths=[1.5 * cm, 3.5 * cm, 3 * cm, 3 * cm, 4 * cm], repeatRows=1)
        cuotas_table.setStyle(TableStyle(cuotas_style))
        elements.append(cuotas_table)

    doc.build(elements)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{_export_filename(purchase)}.pdf"'
    return response
