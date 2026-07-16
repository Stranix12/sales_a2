"""Exportación (PDF/Excel) del reporte de IVA por período.

Mismo estándar visual que ``purchasing.exports`` / ``ExportListMixin`` para
que todos los documentos generados por la aplicación se vean consistentes:
encabezado de tabla ``#343A40`` en blanco/negrita, filas alternas grises,
nombre de archivo con marca de tiempo.
"""
from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

HEADER_FILL = '343A40'
ACCENT = '4B57C4'


def _export_filename(date_from, date_to):
    stamp = timezone.localtime().strftime('%Y%m%d_%H%M')
    return f'Reporte_IVA_{date_from:%Y%m%d}_{date_to:%Y%m%d}_{stamp}'


def iva_report_excel_response(date_from, date_to, filas, totales):
    """Genera el .xlsx del reporte de IVA: resumen del período + detalle por
    factura, listo para adjuntar a una declaración."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Reporte de IVA'

    bold = Font(bold=True)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type='solid')
    center = Alignment(horizontal='center', vertical='center')

    ws.append(['Reporte de IVA'])
    ws.append(['Período', f'{date_from:%d/%m/%Y} — {date_to:%d/%m/%Y}'])
    ws.append(['Generado', timezone.localtime().strftime('%d/%m/%Y %H:%M')])
    for row in range(1, 4):
        ws.cell(row=row, column=1).font = bold
    ws.append([])

    ws.append(['Base imponible 15%', float(totales['base_15'])])
    ws.append(['IVA generado (15%)', float(totales['iva_15'])])
    ws.append(['Base imponible 0%', float(totales['base_0'])])
    ws.append(['Total base imponible', float(totales['total_base'])])
    ws.append(['Total IVA', float(totales['total_iva'])])
    ws.append(['TOTAL', float(totales['total'])])
    for row in range(6, 12):
        ws.cell(row=row, column=1).font = bold
    ws.append([])

    header_row = ws.max_row + 1
    headers = ['N.º Factura', 'Fecha', 'Cliente', 'Base 15%', 'IVA 15%', 'Base 0%', 'Total']
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.font, cell.fill, cell.alignment = header_font, header_fill, center

    for fila in filas:
        inv = fila['invoice']
        ws.append([
            inv.numero_factura or f'#{inv.id}',
            timezone.localtime(inv.invoice_date).strftime('%d/%m/%Y'),
            str(inv.customer),
            float(fila['base_15']), float(fila['iva_15']),
            float(fila['base_0']), float(fila['total']),
        ])

    widths = [18, 12, 28, 12, 12, 12, 12]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{_export_filename(date_from, date_to)}.xlsx"'
    return response


def iva_report_pdf_response(date_from, date_to, filas, totales):
    """Genera el PDF del reporte de IVA: resumen del período + detalle por
    factura."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        topMargin=1.2 * cm, bottomMargin=1.2 * cm,
        leftMargin=1.2 * cm, rightMargin=1.2 * cm,
        title='Reporte de IVA',
    )
    styles = getSampleStyleSheet()
    normal = styles['Normal']
    label_style = ParagraphStyle('lbl', parent=normal, fontName='Helvetica-Bold')

    elements = [
        Paragraph('Reporte de IVA', styles['Title']),
        Paragraph(
            f'Período: {date_from:%d/%m/%Y} — {date_to:%d/%m/%Y} &middot; '
            f"Generado: {timezone.localtime().strftime('%d/%m/%Y %H:%M')}", normal),
        Spacer(1, 0.4 * cm),
        Table([
            [Paragraph('Base imponible 15%:', label_style), f"${totales['base_15']}",
             Paragraph('Base imponible 0%:', label_style), f"${totales['base_0']}"],
            [Paragraph('IVA generado:', label_style), f"${totales['iva_15']}",
             Paragraph('Total base imponible:', label_style), f"${totales['total_base']}"],
            [Paragraph('TOTAL:', label_style), f"${totales['total']}", '', ''],
        ], colWidths=[4.5 * cm, 3.5 * cm, 5 * cm, 3.5 * cm]),
        Spacer(1, 0.6 * cm),
    ]

    data = [['N.º Factura', 'Fecha', 'Cliente', 'Base 15%', 'IVA 15%', 'Base 0%', 'Total']]
    for fila in filas:
        inv = fila['invoice']
        data.append([
            inv.numero_factura or f'#{inv.id}',
            timezone.localtime(inv.invoice_date).strftime('%d/%m/%Y'),
            str(inv.customer),
            f"${fila['base_15']}", f"${fila['iva_15']}", f"${fila['base_0']}", f"${fila['total']}",
        ])
    if not filas:
        data.append(['Sin facturas en este período.', '', '', '', '', '', ''])

    table = Table(data, colWidths=[3.5 * cm, 2.5 * cm, 6 * cm, 2.8 * cm, 2.8 * cm, 2.8 * cm, 2.8 * cm], repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f'#{HEADER_FILL}')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
        ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)

    doc.build(elements)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{_export_filename(date_from, date_to)}.pdf"'
    return response
