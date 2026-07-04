"""Mixins reutilizables para las vistas de la app billing."""
import os
from datetime import datetime, date
from decimal import Decimal
from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Image as RLImage
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


class ExportListMixin:
    """Listado con columnas configurables + exportación a PDF/Excel.

    **Única fuente de configuración de columnas**: la lista ``export_columns``.
    El listado HTML, el PDF y el Excel se construyen a partir de ella, de modo
    que siempre muestran/exportan exactamente lo mismo, en el mismo orden.

    Cada columna es un dict:
        {'key': 'unit_price', 'label': 'Precio', 'field': 'unit_price',
         'default': True, 'type': 'text'}
      - key:     identificador único (usado en el checklist y la sesión)
      - label:   encabezado mostrado
      - field:   ruta para obtener el valor (admite 'brand.name', M2M, props)
      - default: si está visible por defecto
      - type:    'text' (por defecto) o 'image' (se embebe la imagen)

    Columnas visibles:
      - Se persisten en sesión (``columns_session_key``) → se recuerdan al navegar.
      - Se actualizan con ``?cols=key1&cols=key2`` (checklist) y se reinician con
        ``?reset_cols=1``.
      - Mínimo 1 columna; si se envía vacío se conserva la selección previa.

    Exportación: ``?export=pdf`` o ``?export=excel`` (respeta filtros + columnas).
    """
    export_columns = None
    columns_session_key = None
    export_title = None
    export_filename = None

    # ------------------------------------------------------- registro columnas
    def get_columns(self):
        if self.export_columns:
            return [dict(c) for c in self.export_columns]
        # Fallback: todos los campos concretos del modelo.
        return [{'key': f.name, 'label': f.name.replace('_', ' ').title(),
                 'field': f.name, 'default': True, 'type': 'text'}
                for f in self.model._meta.fields]

    def get_columns_session_key(self):
        return self.columns_session_key or f'cols_{self.model._meta.label_lower}'

    def get_visible_column_keys(self):
        """Resuelve las claves visibles desde GET (que actualiza la sesión),
        la sesión, o los valores por defecto. Mantiene el orden del registro."""
        all_keys = [c['key'] for c in self.get_columns()]
        skey = self.get_columns_session_key()
        get = self.request.GET

        if 'reset_cols' in get:
            self.request.session.pop(skey, None)
        elif 'cols' in get:
            chosen = [k for k in all_keys if k in get.getlist('cols')]
            if chosen:  # mínimo 1; si viene vacío se ignora
                self.request.session[skey] = chosen

        stored = self.request.session.get(skey)
        if stored:
            keys = [k for k in all_keys if k in stored]
            if keys:
                return keys
        return [c['key'] for c in self.get_columns() if c.get('default', True)]

    def get_visible_columns(self):
        visible = set(self.get_visible_column_keys())
        return [c for c in self.get_columns() if c['key'] in visible]

    # ------------------------------------------------------------------ config
    def get_export_fields(self):
        return [c['field'] for c in self.get_visible_columns()]

    def get_export_headers(self):
        return [c['label'] for c in self.get_visible_columns()]

    def get_export_title(self):
        if self.export_title:
            return self.export_title
        meta = getattr(self, 'model', None) and self.model._meta
        return str(getattr(meta, 'verbose_name_plural', 'Listado')).title()

    def get_export_filename(self):
        base = self.export_filename or self.get_export_title()
        stamp = timezone.localtime().strftime('%Y%m%d_%H%M')
        return f'{base}_{stamp}'.replace(' ', '_')

    # ------------------------------------------------------------------- datos
    def get_export_value(self, obj, field):
        """Resuelve el valor de un campo soportando relaciones (``brand.name``),
        M2M/relaciones inversas (``suppliers``), callables/properties y booleanos."""
        value = obj
        for part in field.replace('__', '.').split('.'):
            if value is None:
                return ''
            value = getattr(value, part, '')
            # Manager M2M / relación inversa: materializar y terminar.
            if hasattr(value, 'all') and callable(value.all):
                return ', '.join(str(x) for x in value.all())
            if callable(value):
                value = value()
        if isinstance(value, bool):
            return 'Sí' if value else 'No'
        return '' if value is None else value

    def _get_field_file(self, obj, field):
        """Devuelve el FieldFile (imagen) de la ruta indicada, o None."""
        value = obj
        for part in field.replace('__', '.').split('.'):
            if value is None:
                return None
            value = getattr(value, part, None)
        return value or None  # FieldFile vacío es falsy

    # ---------------------------------------------------------------- dispatch
    def get(self, request, *args, **kwargs):
        export = request.GET.get('export')
        if export == 'excel':
            return self.export_excel()
        if export == 'pdf':
            return self.export_pdf()
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        # Querystring sin parámetros de control: conserva SOLO los filtros en
        # paginación y exportaciones (las columnas viven en sesión).
        params = self.request.GET.copy()
        for k in ('page', 'export', 'cols', 'reset_cols'):
            params.pop(k, None)
        ctx['querystring'] = params.urlencode()

        all_cols = self.get_columns()
        visible_keys = self.get_visible_column_keys()
        ctx['all_columns'] = [dict(c, selected=c['key'] in visible_keys) for c in all_cols]
        ctx['visible_columns'] = [c for c in all_cols if c['key'] in visible_keys]
        ctx['visible_count'] = len(visible_keys)
        ctx['total_columns'] = len(all_cols)
        return ctx

    # ------------------------------------------------------------------- Excel
    @staticmethod
    def _excel_safe(val):
        if isinstance(val, Decimal):
            return float(val)
        if isinstance(val, datetime):
            return timezone.localtime(val).replace(tzinfo=None) if val.tzinfo else val
        if isinstance(val, (int, float, str, date)):
            return val
        return str(val)

    def _excel_image(self, ws, row, col_idx, obj, coldef):
        f = self._get_field_file(obj, coldef['field'])
        path = getattr(f, 'path', None) if f else None
        if path and os.path.exists(path):
            try:
                img = XLImage(path)
                img.width = img.height = 46
                ws.add_image(img, ws.cell(row=row, column=col_idx).coordinate)
                return True
            except Exception:
                pass
        ws.cell(row=row, column=col_idx, value='—')
        return False

    def export_excel(self):
        columns = self.get_visible_columns()
        objects = list(self.get_queryset())

        wb = openpyxl.Workbook()
        ws = wb.active
        title = self.get_export_title()
        ws.title = ''.join(c for c in title if c not in '[]:*?/\\')[:31] or 'Listado'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='343A40', end_color='343A40', fill_type='solid')
        center = Alignment(horizontal='center', vertical='center')

        widths = []
        for idx, coldef in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=idx, value=coldef['label'])
            cell.font, cell.fill, cell.alignment = header_font, header_fill, center
            widths.append(len(str(coldef['label'])))

        for r, obj in enumerate(objects, start=2):
            has_image = False
            for idx, coldef in enumerate(columns, start=1):
                if coldef.get('type') == 'image':
                    if self._excel_image(ws, r, idx, obj, coldef):
                        has_image = True
                    widths[idx - 1] = max(widths[idx - 1], 10)
                else:
                    val = self._excel_safe(self.get_export_value(obj, coldef['field']))
                    ws.cell(row=r, column=idx, value=val)
                    widths[idx - 1] = max(widths[idx - 1], len(str(val)))
            if has_image:
                ws.row_dimensions[r].height = 38

        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = min(width + 2, 50)
        ws.freeze_panes = 'A2'

        buffer = BytesIO()
        wb.save(buffer)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{self.get_export_filename()}.xlsx"'
        return response

    # --------------------------------------------------------------------- PDF
    @staticmethod
    def _pdf_safe(val):
        if isinstance(val, datetime):
            val = timezone.localtime(val) if val.tzinfo else val
            return val.strftime('%d/%m/%Y %H:%M')
        if isinstance(val, date):
            return val.strftime('%d/%m/%Y')
        return '' if val is None else str(val)

    def _pdf_cell(self, obj, coldef, body_style, img_size):
        if coldef.get('type') == 'image':
            f = self._get_field_file(obj, coldef['field'])
            path = getattr(f, 'path', None) if f else None
            if path and os.path.exists(path):
                try:
                    return RLImage(path, width=img_size, height=img_size)
                except Exception:
                    pass
            return Paragraph('—', body_style)
        return Paragraph(self._pdf_safe(self.get_export_value(obj, coldef['field'])), body_style)

    def export_pdf(self):
        columns = self.get_visible_columns()
        objects = list(self.get_queryset())
        n = len(columns)

        # Adaptación automática según el nº de columnas.
        many = n > 4
        pagesize = landscape(A4) if many else A4
        font_size = 9 if n <= 4 else 8 if n <= 7 else 7 if n <= 9 else 6

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=pagesize,
            topMargin=1.2 * cm, bottomMargin=1.2 * cm,
            leftMargin=1 * cm, rightMargin=1 * cm,
            title=self.get_export_title(),
        )
        styles = getSampleStyleSheet()
        head_style = ParagraphStyle('eh', parent=styles['Normal'], fontName='Helvetica-Bold',
                                    fontSize=font_size, textColor=colors.white, alignment=1)
        body_style = ParagraphStyle('eb', parent=styles['Normal'],
                                    fontSize=font_size, leading=font_size + 2)

        elements = [
            Paragraph(self.get_export_title(), styles['Title']),
            Paragraph(
                f"Generado: {timezone.localtime().strftime('%d/%m/%Y %H:%M')} &middot; "
                f"Registros: {len(objects)} &middot; Columnas: {n}", styles['Normal']),
            Spacer(1, 0.4 * cm),
        ]

        if objects and columns:
            usable = pagesize[0] - doc.leftMargin - doc.rightMargin
            img_col_w = 1.4 * cm
            n_img = sum(1 for c in columns if c.get('type') == 'image')
            n_text = n - n_img
            text_w = (usable - n_img * img_col_w) / n_text if n_text else usable
            col_widths = [img_col_w if c.get('type') == 'image' else text_w for c in columns]
            img_size = img_col_w - 0.3 * cm

            data = [[Paragraph(str(c['label']), head_style) for c in columns]]
            for obj in objects:
                data.append([self._pdf_cell(obj, c, body_style, img_size) for c in columns])

            table = Table(data, colWidths=col_widths, repeatRows=1)
            table.hAlign = 'CENTER'  # centra la tabla (útil con pocas columnas)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#343A40')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ]))
            elements.append(table)
        else:
            elements.append(Paragraph('Sin registros para los filtros aplicados.', styles['Normal']))

        doc.build(elements)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{self.get_export_filename()}.pdf"'
        return response
